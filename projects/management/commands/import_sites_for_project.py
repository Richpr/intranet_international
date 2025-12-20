# projects/management/commands/import_sites_for_project.py

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from datetime import date, timedelta
from django.utils import timezone
from projects.models import (
    Site, 
    Project, 
    SiteType, 
    InstallationType, 
    Task, 
    TaskType, 
    TaskResultType, 
    SiteRadioConfiguration, 
    RadioType,              
    AntennaType, 
    BBMLType, 
    EnclosureType, 
)
from users.models import CustomUser 


# Fonction utilitaire pour trouver les objets par nom/ID (Recherche Tolérante)
def get_related_object(Model, identifier, lookup_field='name', required=False):
    if not identifier:
        return None
        
    identifier_cleaned = str(identifier).strip() 
    
    # Nettoyage pour la tolérance (élimine l'encodage problématique et met en MAJUSCULES)
    try:
        identifier_cleaned = identifier_cleaned.encode('ascii', 'ignore').decode('ascii').upper()
    except Exception:
        pass 
    
    try:
        filters = {f'{lookup_field}__iexact': identifier_cleaned}
        
        return Model.objects.get(**filters)
        
    except Model.DoesNotExist:
        if required:
            raise CommandError(f"Clé étrangère manquante: Objet {Model.__name__} avec '{identifier_cleaned}' non trouvé.")
        return None
    except Exception as e:
        raise CommandError(f"Erreur lors de la recherche de l'objet {Model.__name__} ('{identifier_cleaned}'): {e}")


class Command(BaseCommand):
    help = "Importe des Sites pour un Project spécifique à partir d'un fichier Excel (XLSX), crée la configuration radio et 7 tâches terminées."

    def add_arguments(self, parser):
        parser.add_argument('project_pk', type=int, help="L'ID (PK) du Projet parent.")
        parser.add_argument('file_path', type=str, help='Le chemin complet vers le fichier Excel.')

    @transaction.atomic
    def handle(self, *args, **options):
        project_pk = options['project_pk']
        file_path = options['file_path']
        sites_a_creer = []
        lignes_traitees = 0
        radio_data_pour_etape_suivante = [] 

        creator_user = CustomUser.objects.filter(is_superuser=True).first()
        if not creator_user:
            creator_user = CustomUser.objects.filter(is_active=True).first()
            if not creator_user:
                self.stdout.write(self.style.WARNING("⚠️ Aucun utilisateur actif trouvé. created_by sera None."))
        
        # 1. Vérifier et récupérer le Project cible
        try:
            target_project = Project.objects.get(pk=project_pk)
            self.stdout.write(self.style.SUCCESS(f"Projet cible trouvé: {target_project.name}"))
        except Project.DoesNotExist:
            raise CommandError(f"Le Projet avec l'ID {project_pk} n'existe pas.")

        self.stdout.write(f"Chargement du fichier Excel : {file_path}...")

        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # --- DÉFINITION DE LA CORRESPONDANCE DES COLONNES (INDICES) ---
            COL_ID_CLIENT = 0     
            COL_NAME = 1          
            COL_SITE_TYPE = 2     
            COL_INSTALL_TYPE = 3  
            COL_TEAM_LEAD = 4     
            COL_START_DATE = 5    
            COL_RADIO_MODEL = 6   
            COL_RADIO_QTY = 7     
            COL_ANTENNA_TYPE = 8  
            COL_BB_ML = 9         
            COL_ENCLOSURE_TYPE = 10 
            # ...

            # Itération sur les lignes à partir de la deuxième
            for row in sheet.iter_rows(min_row=2, values_only=True):
                lignes_traitees += 1
                row_len = len(row) 

                if not row[COL_ID_CLIENT] or not row[COL_NAME]:
                    self.stdout.write(self.style.WARNING(f"⚠️ Ligne {lignes_traitees}: ID Client ou Nom manquant. Ligne ignorée."))
                    continue

                site_id_client = str(row[COL_ID_CLIENT]).strip()
                site_name = str(row[COL_NAME]).strip()
                
                # --- 2. Gérer les dépendances (Clés Étrangères) ---
                try:
                    # Remplissage SiteType (Required=False car null=True)
                    site_type_instance = get_related_object(SiteType, row[COL_SITE_TYPE], required=False)
                    
                    # 🚨 CORRECTION CRITIQUE : Gestion robuste du Type d'Installation
                    install_type_excel_value = row[COL_INSTALL_TYPE]
                    install_type_instance = None
                    
                    if install_type_excel_value:
                        # Si la cellule n'est PAS vide, on force la recherche
                        install_type_instance = get_related_object(InstallationType, install_type_excel_value, required=True) 
                    else:
                        # 🚨 CRÉATION D'UN TYPE PAR DÉFAUT SI AUCUN N'EST FOURNI
                        install_type_instance, created = InstallationType.objects.get_or_create(
                            name="Non spécifié",
                            defaults={'is_active': True}  # ✅ CORRECTION : pas de champ 'code'
                        )
                        if created:
                            self.stdout.write(self.style.SUCCESS("✅ Type d'installation par défaut créé: 'Non spécifié'"))

                    team_lead_username = str(row[COL_TEAM_LEAD]).strip() if row[COL_TEAM_LEAD] else None
                    team_lead_instance = get_related_object(CustomUser, team_lead_username, lookup_field='username', required=False) 

                    # 🚨 CORRECTION : Gestion robuste des dates
                    start_date_value = row[COL_START_DATE]
                    final_start_date = date.today()

                    try:
                        if isinstance(start_date_value, str) and start_date_value.strip():
                            final_start_date = date.fromisoformat(str(start_date_value).strip())
                        elif isinstance(start_date_value, date):
                            final_start_date = start_date_value
                        elif start_date_value:
                            # Gestion des dates Excel (nombre de jours depuis 1900)
                            final_start_date = date(1900, 1, 1) + timedelta(days=int(start_date_value) - 2)
                    except (ValueError, TypeError, AttributeError) as e:
                        self.stdout.write(self.style.WARNING(f"⚠️ Date invalide à la ligne {lignes_traitees}, utilisation de la date du jour. Erreur: {e}"))
                        final_start_date = date.today()

                    # Champs techniques optionnels (Required=False)
                    antenna_type_name = row[COL_ANTENNA_TYPE] if row_len > COL_ANTENNA_TYPE else None
                    bb_ml_name = row[COL_BB_ML] if row_len > COL_BB_ML else None
                    enclosure_type_name = row[COL_ENCLOSURE_TYPE] if row_len > COL_ENCLOSURE_TYPE else None
                    
                    antenna_type_instance = get_related_object(AntennaType, antenna_type_name, required=False)
                    bb_ml_instance = get_related_object(BBMLType, bb_ml_name, required=False)
                    enclosure_type_instance = get_related_object(EnclosureType, enclosure_type_name, required=False)
                    
                    # Données Radio pour l'étape post-création
                    radio_model_name = str(row[COL_RADIO_MODEL]).strip() if row[COL_RADIO_MODEL] else None
                    radio_qty_value = int(row[COL_RADIO_QTY]) if row[COL_RADIO_QTY] else 0

                    if radio_model_name and radio_qty_value > 0:
                        radio_data_pour_etape_suivante.append({
                            'site_index': len(sites_a_creer),
                            'radio_model_name': radio_model_name,
                            'quantity': radio_qty_value
                        })

                except CommandError as e:
                    # Arrête ici pour forcer l'utilisateur à corriger l'objet manquant
                    raise e
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"❌ Erreur de données non gérée à la ligne {lignes_traitees}: {e}"))
                    continue


                # --- 3. Créer l'objet Site ---
                site = Site(
                    project=target_project,
                    site_id_client=site_id_client,
                    name=site_name,
                    start_date=final_start_date,
                    
                    site_type=site_type_instance,
                    installation_type=install_type_instance, # 🚨 MAINTENANT TOUJOURS NON-NULL
                    team_lead=team_lead_instance,
                    
                    antenna_type=antenna_type_instance,
                    bb_ml=bb_ml_instance,
                    enclosure_type=enclosure_type_instance,
                    
                    created_by=creator_user
                )
                sites_a_creer.append(site)

            # --- 4. Insertion des Sites en masse (Le Site reçoit sa PK) ---
            if sites_a_creer:
                Site.objects.bulk_create(sites_a_creer)
                self.stdout.write(self.style.SUCCESS(f"✅ Importation des sites terminée : {len(sites_a_creer)} sites insérés."))
            else:
                self.stdout.write(self.style.WARNING("Aucun site valide n'a été trouvé dans le fichier pour l'importation."))
                return

            # ==========================================================
            # 5. CRÉATION DES 7 TÂCHES INITIALES TERMINÉES

            TASK_CODES_TO_COMPLETE = ['CLEANUP', 'ANTENNA_INSTALL', 'QA_PHOTOS', 'EHS_PRE', 'ATP', 'SRS', 'IMK'] 
            SUCCESS_RESULT_CODE = 'DONE' 
            
            assigned_user = creator_user or CustomUser.objects.filter(is_active=True).first()
            if not assigned_user:
                raise CommandError("Impossible d'assigner les tâches: Aucun utilisateur actif trouvé.")

            result_type_instance = get_related_object(TaskResultType, SUCCESS_RESULT_CODE, lookup_field='code', required=True)
            current_date = date.today()
            tasks_a_creer = []
            
            task_type_instances = {
                t.code: t for t in TaskType.objects.filter(code__in=TASK_CODES_TO_COMPLETE)
            }

            if len(task_type_instances) != len(TASK_CODES_TO_COMPLETE):
                 missing = set(TASK_CODES_TO_COMPLETE) - set(task_type_instances.keys())
                 raise CommandError(f"❌ Erreur critique : Les TaskType suivants sont manquants en base de données : {', '.join(missing)}. Veuillez les créer dans l'admin.")
            
            for site in sites_a_creer:
                for code in TASK_CODES_TO_COMPLETE:
                    task_type = task_type_instances[code]

                    tasks_a_creer.append(
                        Task(
                            site=site,
                            task_type=task_type,
                            description=f"Tâche initiale '{task_type.name}' - Importation de Site",
                            assigned_to=assigned_user, 
                            due_date=current_date, 
                            
                            status='COMPLETED',
                            progress_percentage=100,
                            result_type=result_type_instance,
                            completion_date=timezone.now(),  # ✅ CORRECTION : utilisation de timezone
                            
                            created_by=creator_user
                        )
                    )

            Task.objects.bulk_create(tasks_a_creer)
            self.stdout.write(self.style.SUCCESS(f"✅ {len(tasks_a_creer)} tâches ({len(TASK_CODES_TO_COMPLETE)} types) créées et marquées comme Complétées (100%)."))
            
            # ==========================================================
            # 6. CRÉATION DES CONFIGURATIONS RADIO
            
            radio_configs_a_creer = []
            
            for data in radio_data_pour_etape_suivante:
                
                site_instance = sites_a_creer[data['site_index']] 
                
                try:
                    radio_type_instance = get_related_object(RadioType, data['radio_model_name'], required=True)
                
                    radio_configs_a_creer.append(
                        SiteRadioConfiguration(
                            site=site_instance,
                            radio_type=radio_type_instance,
                            quantity=data['quantity']
                        )
                    )
                
                except CommandError as e:
                    self.stdout.write(self.style.ERROR(f"❌ Radio manquée pour le site {site_instance.site_id_client}: {e}"))
                    continue


            if radio_configs_a_creer:
                SiteRadioConfiguration.objects.bulk_create(radio_configs_a_creer)
                self.stdout.write(self.style.SUCCESS(f"✅ {len(radio_configs_a_creer)} configurations radio créées avec succès (les manquantes ont été ignorées)."))
            
            # 7. Mise à jour de la progression du projet
            for site in sites_a_creer:
                site.update_progress()
            
            self.stdout.write(self.style.SUCCESS("✅ Mise à jour de la progression des sites et du projet effectuée."))

        except CommandError as e:
            # 🚨 Si l'erreur est levée ici, l'utilisateur DOIT corriger la donnée dans l'admin/Excel
            raise e
        except FileNotFoundError:
            raise CommandError(f'Le fichier "{file_path}" n\'existe pas.')
        except Exception as e:
            raise CommandError(f"Une erreur inattendue s'est produite: {e}")