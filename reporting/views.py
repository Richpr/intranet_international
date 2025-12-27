from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from projects.models import Site, TransmissionLink, Project  # 👈 AJOUT DE PROJECT
from users.models import Country
from logistique.models import Vehicule
from inventaire.models import Equipement
from django.db.models import Sum, F, Q, Count
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
from datetime import date
from .utils import get_project_performance_by_year, get_site_completion_rate_by_year, get_site_profitability_by_year

@login_required
def performance_annuelle_view(request):
    countries = Country.objects.filter(is_active=True)
    selected_country_id = request.GET.get('country')

    project_performance_qs = get_project_performance_by_year(selected_country_id)
    site_completion_qs = get_site_completion_rate_by_year(selected_country_id)
    site_profitability_qs = get_site_profitability_by_year(selected_country_id)

    project_performance = [
        {
            'year': item['year'],
            'total_budget': float(item['total_budget']),
            'total_progress': float(item['total_progress']),
        }
        for item in project_performance_qs
    ]

    context = {
        'project_performance': project_performance,
        'site_completion': list(site_completion_qs),
        'site_profitability': list(site_profitability_qs),
        'countries': countries,
        'selected_country_id': selected_country_id,
    }

    return render(request, 'reporting/performance_annuelle.html', context)


# =================================================================
# VUES RENTABILITÉ
# =================================================================

@login_required
def site_profitability_report_view(request):
    countries = Country.objects.filter(is_active=True)
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')

    sites_qs = Site.objects.all().select_related('project__country')
    projects_qs = Project.objects.filter(is_active=True)

    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
        projects_qs = projects_qs.filter(country_id=selected_country_id)

    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)

    sites_with_profit = sites_qs.annotate(
        total_expenses=Sum('depenses__montant'),
        profit=F('prix_facturation') - F('total_expenses')
    )

    context = {
        'sites': sites_with_profit,
        'countries': countries,
        'projects': projects_qs,
        'selected_country_id': selected_country_id,
        'selected_project_id': selected_project_id,
    }
    return render(request, 'reporting/site_profitability_report.html', context)

@login_required
def cost_per_vehicle_report_view(request):
    vehicules_qs = Vehicule.objects.all()

    vehicules_with_cost = vehicules_qs.annotate(
        total_cost=Sum('depenses__montant', filter=Q(depenses__categorie__in=['CARBURANT', 'REPARATION_VEHICULE']))
    )

    context = {
        'vehicules': vehicules_with_cost,
    }
    return render(request, 'reporting/cost_per_vehicle_report.html', context)

@login_required
def inventory_status_report_view(request):
    equipments_qs = Equipement.objects.values('statut').annotate(count=Count('statut'))

    context = {
        'equipments': equipments_qs,
    }
    return render(request, 'reporting/inventory_status_report.html', context)

# =================================================================
# VUES RAN
# =================================================================

@login_required



def ran_site_list_view(request):

    countries = Country.objects.filter(is_active=True)
    
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    projects_qs = Project.objects.filter(is_active=True, project_type__is_transmission=False)
    
    # 💡 CORRECTION : select_related uniquement pour les ForeignKey directes.
    # On ajoute 'prefetch_related' pour les inspections et les tâches afin d'optimiser le template.
    sites_qs = Site.objects.filter(project__project_type__is_transmission=False).select_related(
        'project__country', 
        'departement', 
        'site_type', 
        'antenna_type', 
        'enclosure_type', 
        'bb_ml'
    ).prefetch_related(
        'radio_configurations__radio_type',
        'inspections',
        'tasks'
    )

    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
        projects_qs = projects_qs.filter(country_id=selected_country_id)
        
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)

    context = {
        'sites': sites_qs,
        'countries': countries,
        'projects': projects_qs,
        'selected_country_id': selected_country_id,
        'selected_project_id': selected_project_id,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'years': range(date.today().year + 1, date.today().year - 5, -1),
        'months': range(1, 13),
    }
    return render(request, 'reporting/ran_site_list.html', context)

@login_required
def ran_site_list_pdf(request):

    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    sites_qs = Site.objects.filter(project__project_type__is_transmission=False).select_related(
        'project__country', 'departement', 'site_type', 
        'antenna_type', 'enclosure_type', 'bb_ml'
    ).prefetch_related('radio_configurations__radio_type')

    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)

    html_string = render_to_string('reporting/ran_site_list_pdf.html', {'sites': sites_qs})
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ran_sites_report.pdf"'
    return response

@login_required

def ran_site_list_excel(request):
    # 1. RÉCUPÉRATION DES FILTRES (Identiques à ta vue de liste)
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    # 2. QUERYSET OPTIMISÉ (Important pour ne pas faire ramer le serveur)
    sites_qs = Site.objects.filter(
        project__project_type__is_transmission=False
    ).select_related(
        'project__country', 
        'departement', 
        'site_type', 
        'antenna_type', 
        'enclosure_type', 
        'bb_ml'
    ).prefetch_related(
        'radio_configurations__radio_type',
        'tasks__task_type',
        'inspections'
    ).order_by('name')

    # Application des filtres
    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)

    # 3. CRÉATION DU WORKBOOK EXCEL
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Rapport Déploiement RAN'

    # Headers complets (Identification + Technique + Tes 9 Tâches)
    headers = [
        "Site Name", "Site Id", "Pays", "Departement", "Commune", "Site Type", 
        "Project Name", "Année", "Mois", "Radio Type", "Antenna Type", 
        "Enclosure", "BB/ML", "INSTALLATION", "INTEGRATION", "EHS 1", 
        "EHS 2", "IMK", "SRS", "QA (RESULT)", "ATP", "COMMENT"
    ]
    sheet.append(headers)

    # 4. REMPLISSAGE DES DONNÉES
    for site in sites_qs:
        # On récupère toutes les tâches du site une seule fois
        tasks = site.tasks.all()
        
        # Fonction interne pour extraire le statut d'une tâche par son nom
        def get_task_status(name_to_find):
            for t in tasks:
                if name_to_find.upper() in t.task_type.name.upper():
                    return t.get_status_display()
            return "-"

        # Préparation de la ligne
        row = [
            site.name,
            site.site_id_client,
            site.project.country.name,
            site.departement.name if site.departement else "-",
            site.location if site.location else "-",
            site.site_type.name if site.site_type else "-",
            site.project.name,
            site.start_date.year if site.start_date else "-",
            site.start_date.strftime('%B') if site.start_date else "-",
            ", ".join([r.radio_type.name for r in site.radio_configurations.all()]),
            site.antenna_type.name if site.antenna_type else "-",
            site.enclosure_type.name if site.enclosure_type else "-",
            site.bb_ml.name if site.bb_ml else "-",
            # Les 9 colonnes demandées
            get_task_status("INSTALLATION"),
            get_task_status("INTEGRATION"),
            get_task_status("EHS 1"),
            get_task_status("EHS 2"),
            get_task_status("IMK"),
            get_task_status("SRS"),
            # QA basé sur l'inspection
            site.inspections.first().get_resultat_inspection_display() if site.inspections.exists() else "-",
            get_task_status("ATP"),
            site.comment if site.comment else "-"
        ]
        sheet.append(row)

    # 5. MISE EN FORME BASIQUE (Optionnel : figer la première ligne)
    sheet.freeze_panes = 'A2'
    
    # 6. GÉNÉRATION DE LA RÉPONSE
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Rapport_RAN_{date.today().strftime('%d_%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response
# =================================================================
# VUES TRANSMISSION
# =================================================================

@login_required
def transmission_site_list_view(request):
    countries = Country.objects.filter(is_active=True)
    
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    projects_qs = Project.objects.filter(is_active=True, project_type__is_transmission=True)
    links_qs = TransmissionLink.objects.all().select_related(
        'site_a__project__country', 'site_a__departement', 'site_a__antenna_type', 
        'site_a__enclosure_type', 'site_a__bb_ml', 'site_b'
    ).prefetch_related('site_a__radio_configurations__radio_type')

    if selected_country_id:
        links_qs = links_qs.filter(site_a__project__country_id=selected_country_id)
        projects_qs = projects_qs.filter(country_id=selected_country_id)
        
    if selected_project_id:
        links_qs = links_qs.filter(site_a__project_id=selected_project_id)
    if selected_year:
        links_qs = links_qs.filter(created_at__year=selected_year) # Filtre sur link.created_at
    if selected_month:
        links_qs = links_qs.filter(created_at__month=selected_month) # Filtre sur link.created_at
        
    context = {
        'links': links_qs,
        'countries': countries,
        'projects': projects_qs,
        'selected_country_id': selected_country_id,
        'selected_project_id': selected_project_id,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'years': range(date.today().year + 1, date.today().year - 5, -1),
        'months': range(1, 13),
    }
    return render(request, 'reporting/transmission_site_list.html', context)

@login_required
def transmission_site_list_pdf(request):
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    links_qs = TransmissionLink.objects.all().select_related(
        'site_a__project__country', 'site_a__departement', 'site_a__antenna_type', 
        'site_a__enclosure_type', 'site_a__bb_ml', 'site_b'
    ).prefetch_related('site_a__radio_configurations__radio_type')

    if selected_country_id:
        links_qs = links_qs.filter(site_a__project__country_id=selected_country_id)
    if selected_project_id:
        links_qs = links_qs.filter(site_a__project_id=selected_project_id)
    if selected_year:
        links_qs = links_qs.filter(created_at__year=selected_year)
    if selected_month:
        links_qs = links_qs.filter(created_at__month=selected_month)

    html_string = render_to_string('reporting/transmission_site_list_pdf.html', {'links': links_qs})
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transmission_sites_report.pdf"'
    return response

@login_required
def transmission_site_list_excel(request):
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    links_qs = TransmissionLink.objects.all().select_related(
        'site_a__project__country', 'site_a__departement', 'site_a__antenna_type', 
        'site_a__enclosure_type', 'site_a__bb_ml', 'site_b'
    ).prefetch_related('site_a__radio_configurations__radio_type')
    
    if selected_country_id:
        links_qs = links_qs.filter(site_a__project__country_id=selected_country_id)
    if selected_project_id:
        links_qs = links_qs.filter(site_a__project_id=selected_project_id)
    if selected_year:
        links_qs = links_qs.filter(created_at__year=selected_year)
    if selected_month:
        links_qs = links_qs.filter(created_at__month=selected_month)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Transmission Sites'
    headers = [
        "Link", "Site A", "Site B", "Project Name", "Pays", "Departement", "Commune", 
        "Annee", "Mois", "Type de radio", "Type d'antenne", "Type d'enclosure", 
        "Indoor equipement (BB/ML)", "INSTALLATION", "ALLIGNEMENT", "EHS SITE A", 
        "EHS SITE B", "QA STE A", "QA SITE B", "QA STATUT SITE A", "QA STATUS SITE B", 
        "ATP STATUT SITE A", "ATP STATUT SITE B", "COMMENT"
    ]
    sheet.append(headers)
    for link in links_qs:
        radios = ", ".join([f"{radio.radio_type.name} ({radio.quantity})" for radio in link.site_a.radio_configurations.all()])
        row = [
            link.link_id, link.site_a.name, link.site_b.name,
            link.site_a.project.name, link.site_a.project.country.name,
            link.site_a.departement.name if link.site_a.departement else "",
            link.site_a.location, link.created_at.year, link.created_at.month,
            radios, link.site_a.antenna_type.name if link.site_a.antenna_type else "",
            link.site_a.enclosure_type.name if link.site_a.enclosure_type else "",
            link.site_a.bb_ml.name if link.site_a.bb_ml else "",
            str(link.site_a.installation_status), "N/A",
            str(link.site_a.ehs_status), str(link.site_b.ehs_status),
            str(link.site_a.qa_result), str(link.site_b.qa_result),
            str(link.site_a.qa_result), str(link.site_b.qa_result),
            str(link.site_a.atp_status), str(link.site_b.atp_status),
            link.site_a.comment,
        ]
        sheet.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transmission_sites_report.xlsx"'
    workbook.save(response)
    return response

# =================================================================
# VUES SURVEY
# =================================================================

@login_required
def survey_site_list_view(request):
    countries = Country.objects.filter(is_active=True)
    
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    projects_qs = Project.objects.filter(is_active=True, project_type__name='SURVEY')
    sites_qs = Site.objects.filter(project__project_type__name='SURVEY').select_related(
        'project__country', 'departement', 'site_type'
    )
    
    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
        projects_qs = projects_qs.filter(country_id=selected_country_id)
        
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)

    context = {
        'sites': sites_qs,
        'countries': countries,
        'projects': projects_qs,
        'selected_country_id': selected_country_id,
        'selected_project_id': selected_project_id,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'years': range(date.today().year + 1, date.today().year - 5, -1),
        'months': range(1, 13),
    }
    return render(request, 'reporting/survey_site_list.html', context)

@login_required
def survey_site_list_pdf(request):
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    sites_qs = Site.objects.filter(project__project_type__name='SURVEY').select_related(
        'project__country', 'departement', 'site_type'
    )
    
    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)

    html_string = render_to_string('reporting/survey_site_list_pdf.html', {'sites': sites_qs})
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="survey_sites_report.pdf"'
    return response

@login_required
def survey_site_list_excel(request):
    selected_country_id = request.GET.get('country')
    selected_project_id = request.GET.get('project')
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    sites_qs = Site.objects.filter(project__project_type__name='SURVEY').select_related(
        'project__country', 'departement', 'site_type'
    )
    
    if selected_country_id:
        sites_qs = sites_qs.filter(project__country_id=selected_country_id)
    if selected_project_id:
        sites_qs = sites_qs.filter(project_id=selected_project_id)
    if selected_year:
        sites_qs = sites_qs.filter(start_date__year=selected_year)
    if selected_month:
        sites_qs = sites_qs.filter(start_date__month=selected_month)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Survey Sites'
    headers = [
        "Nom du site", "Id du site", "Pays", "Departement", "Commune", "Type de site", 
        "Project Name", "Annee", "Mois", "SURVEY STATUT", "RAPPORT STATUT", "COMMENT"
    ]
    sheet.append(headers)
    for site in sites_qs:
        row = [
            site.name, site.site_id_client, site.project.country.name,
            site.departement.name if site.departement else "",
            site.location, site.site_type.name if site.site_type else "",
            site.project.name,
            site.start_date.year if site.start_date else "",
            site.start_date.month if site.start_date else "",
            "N/A", "N/A", site.comment,
        ]
        sheet.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="survey_sites_report.xlsx"'
    workbook.save(response)
    return response